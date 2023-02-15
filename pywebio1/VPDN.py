import numpy as np
import openpyxl
import pandas as pd
# from matplotlib import pyplot as plt
from openpyxl.reader.excel import *
from pywebio.input import *
from pywebio.output import *
from pywebio.session import *

from pywebio1.PriceCal import *
from pywebio1.UPF import lims_file


# 显示价格详情
def VPDNPrice():
    with use_scope('scope_top', clear=False):
        put_buttons(['价格详情', '价格计算器', '返回主页'],
                    onclick=[lambda: go_app('VPDNPrice', new_window=False),
                             lambda: go_app('VPDNBaoBiao', new_window=False),
                             lambda: go_app('index', new_window=False)])
    with use_scope('scope1', clear=True):
        TableWidth = 'width:1300px'
        put_text('一、定制流量价格参考表').style('font-size:20px;font-weight:700')
        workbook = load_workbook(filename='./pywebio1/价格参考表/5G_VPDN价格参考表.xlsx')
        DingZhiLiuLiang1 = workbook["定制流量1"]

        put_table(
            [
                [span(put_text('定制流量').style(TableWidth), col=12)],
                [span(put_text(DingZhiLiuLiang1['A1'].value), col=2),
                 span(put_text(DingZhiLiuLiang1['C1'].value), col=6),
                 span(put_text(DingZhiLiuLiang1['I1'].value), col=2),
                 span(put_text(DingZhiLiuLiang1['K1'].value), col=2)],
                [put_text(DingZhiLiuLiang1['A2'].value), put_text(DingZhiLiuLiang1['B2'].value),
                 span(put_text(DingZhiLiuLiang1['C2'].value), col=3),
                 span(put_text(DingZhiLiuLiang1['F2'].value), col=3),
                 put_text(DingZhiLiuLiang1['I2'].value), put_text(DingZhiLiuLiang1['J2'].value),
                 put_text(DingZhiLiuLiang1['K2'].value),
                 put_text(DingZhiLiuLiang1['L2'].value)],
                [x for x in info(DingZhiLiuLiang1, 3)],
                [x for x in info(DingZhiLiuLiang1, 4)],
                [x for x in info(DingZhiLiuLiang1, 5)],
                [x for x in info(DingZhiLiuLiang1, 6)],
                [x for x in info(DingZhiLiuLiang1, 7)],
                [x for x in info(DingZhiLiuLiang1, 8)],
                [x for x in info(DingZhiLiuLiang1, 9)],
                [x for x in info(DingZhiLiuLiang1, 10)],
                [x for x in info(DingZhiLiuLiang1, 11)],
                [x for x in info(DingZhiLiuLiang1, 12)],
                [x for x in info(DingZhiLiuLiang1, 13)],
                [x for x in info(DingZhiLiuLiang1, 14)],
                [x for x in info(DingZhiLiuLiang1, 15)],
                [x for x in info(DingZhiLiuLiang1, 16)]
            ]
        ).style('text-align:center;text-align-last:center;table-layout:fixed;word-wrap:break-word')

    with use_scope('scope3', clear=True):
        TableWidth = 'width:790px'
        HalfTableWidth = 'width:395px'
        put_text('二、业务隔离价格参考表').style('font-size:20px;font-weight:700')
        workbook = load_workbook(filename='./pywebio1/价格参考表/UPF价格参考表.xlsx')
        YeWuGeLi = workbook["业务隔离"]
        put_table(
            [
                [span(put_text(YeWuGeLi['A1'].value).style(TableWidth), col=2)],
                [put_text(x).style(HalfTableWidth) for x in info(YeWuGeLi, 2)],
                [put_text(x).style(HalfTableWidth) for x in info(YeWuGeLi, 3)],
            ]
        ).style('text-align:center;text-align-last:center;table-layout:fixed;word-wrap:break-word')

    with use_scope('scope4', clear=True):
        put_text('三、固定入网专线价格参考表').style('font-size:20px;font-weight:700')
        workbook = load_workbook(filename='./pywebio1/价格参考表/5G_VPDN价格参考表.xlsx')
        RuWangZhuanXian = workbook["固定入网专线"]
        TableWidth = 'width:820px'
        put_table(
            [
                [span(put_text(RuWangZhuanXian['A1'].value).style(TableWidth), col=8)],
                [span(put_text(RuWangZhuanXian['A2'].value), col=4), span(put_text(RuWangZhuanXian['E2'].value), col=2),
                 span(put_text(RuWangZhuanXian['G2'].value), col=2)],
                [x for x in info(RuWangZhuanXian, 3)], [x for x in info(RuWangZhuanXian, 4)],
                [x for x in info(RuWangZhuanXian, 5)], [x for x in info(RuWangZhuanXian, 6)],
                [x for x in info(RuWangZhuanXian, 7)], [x for x in info(RuWangZhuanXian, 8)],
                [x for x in info(RuWangZhuanXian, 9)], [x for x in info(RuWangZhuanXian, 10)],
                [x for x in info(RuWangZhuanXian, 11)], [x for x in info(RuWangZhuanXian, 12)],
                [x for x in info(RuWangZhuanXian, 13)]
            ]
        ).style('text-align:center;text-align-last:center;table-layout:fixed;word-wrap:break-word')


def info(table, row):
    # 返回一个列表
    infoList = []
    for x in table[row]:
        if (pd.isnull(x.value) == False):
            infoList.append(x.value)
        else:
            infoList.append('')
    return infoList


# 显示报表信息
def VPDNBaoBiao():
    # 数据导入
    # 解决中文乱码
    # plt.rcParams["font.sans-serif"] = ["SimHei"]  # 设置字体
    # plt.rcParams["axes.unicode_minus"] = False  # 该语句解决图像中的“-”负号的乱码问题
    # xlsx表格数据导入
    RenLian = pd.read_excel('./pywebio1/数据表/人联卡.xlsx')
    WuLian_Month = pd.read_excel('./pywebio1/数据表/物联卡(月包).xlsx')
    WuLian_Year = pd.read_excel('./pywebio1/数据表/物联卡(年包).xlsx')
    LiuLiangChi = pd.read_excel('./pywebio1/数据表/流量池.xlsx')
    HaoKa = pd.read_excel('./pywebio1/数据表/定制号卡.xlsx')
    YeWuGeLi = pd.read_excel('./pywebio1/数据表/业务隔离.xlsx')
    RuWangZhuanXian = pd.read_excel('./pywebio1/数据表/固定入网专线.xlsx')
    WangYuan = pd.read_excel('./pywebio1/数据表/网元定制.xlsx')
    IPRAN = pd.read_excel('./pywebio1/数据表/IPRAN.xlsx')

    # 数据获取
    # b左侧类型 p右侧具体费用
    b1 = np.array(RenLian['月流量/通话'])
    p1 = np.array(RenLian['月资费'])
    b2 = np.array(WuLian_Month['国内流量'])
    b5 = np.array(HaoKa['卡品'])
    b71 = np.array(RuWangZhuanXian['5G云网UPF带宽'])
    b72 = np.array(RuWangZhuanXian['5G定制网STN带宽'])
    b81 = np.array(WangYuan['方案说明'])
    b82 = np.array(WangYuan['协议期'])
    b10 = np.array(IPRAN.iloc[3:, 0])

    b1 = list(b1)
    b2 = list(b2)
    b5 = list(b5)
    b71 = list(b71)
    b72 = list(b72)
    b81 = list(b81)
    b82 = list(b82)
    b10 = list(b10)

    # 复选框的选项
    option_list1 = b1
    option_list2 = b2
    option_list5 = b5
    option_list71 = b71
    option_list72 = b72
    option_list81 = b81
    option_list81 = [x for x in option_list81 if pd.isnull(x) == False]
    option_list82 = b82
    option_list82 = [x for x in option_list82 if pd.isnull(x) == False]
    option_list10 = b10

    with use_scope('scope_top', clear=False):
        put_buttons(['价格详情', '价格计算器', '返回主页'],
                    onclick=[lambda: go_app('VPDNPrice', new_window=False),
                             lambda: go_app('VPDNBaoBiao', new_window=False),
                             lambda: go_app('index', new_window=False)])

    # 需要输入信息的表格
    with use_scope('scope1', clear=True):
        put_text('5G VPDN').style('font-size:20px;font-weight:700;margin:10px')
        put_table(
            [['', '名称', '类型', '规格', '数量', '单位', '税率', '折扣率', '一次性服务费（元）'],
             [span('定制流量', row=5)],
             ['人联卡', '5G畅享套餐', put_select('select1', options=option_list1).style('width:150px'),
              put_input('Num1', type=NUMBER),
              '张', '6%',
              put_row([put_input('Discount1', type=NUMBER), put_text('%').style('margin-top: 5px')], size='80% 20%'),
              put_input('Fee1', type=NUMBER)],
             ['物联卡', put_select('select2', options=['月包(通用)', '月包(定向)', '年包(通用)', '年包(定向)']),
              put_select('select3', options=option_list2), put_input('Num2', type=NUMBER),
              '张', '6%',
              put_row([put_input('Discount2', type=NUMBER), put_text('%').style('margin-top: 5px')], size='80% 20%'),
              put_input('Fee2', type=NUMBER)],
             ['流量池', put_select('select4', options=['定向流量', '通用流量']), '-', put_input('Num3', type=NUMBER),
              'GB', '6%',
              put_row([put_input('Discount3', type=NUMBER), put_text('%').style('margin-top: 5px')], size='80% 20%'),
              put_input('Fee3', type=NUMBER)],
             ['定制号卡', put_select('select5', options=option_list5), '-', put_input('Num5', type=NUMBER),
              '张', '一半6%,一半13%',
              put_row([put_input('Discount5', type=NUMBER), put_text('%').style('margin-top: 5px')], size='80% 20%'),
              '-'],
             [span('业务隔离', row=3)],
             ['定制DNN', '专网加密隧道服务', '-', put_input('Num6', type=NUMBER),
              '线', '6%',
              put_row([put_input('Discount6', type=NUMBER), put_text('%').style('margin-top: 5px')], size='80% 20%'),
              put_input('Fee6', type=NUMBER)],
             ['无线VPDN群', '-', '-', put_input('Num7', type=NUMBER),
              '线', '6%',
              put_row([put_input('Discount7', type=NUMBER), put_text('%').style('margin-top: 5px')], size='80% 20%'),
              put_input('Fee7', type=NUMBER)],
             [span('固定入网专线', row=2)],
             ['5G定制网IPRAN专线', put_select('select11', options=option_list10), '-', put_input('Num11', type=NUMBER),
              '线', '6%',
              put_row([put_input('Discount11', type=NUMBER), put_text('%').style('margin-top: 5px')], size='80% 20%'),
              put_input('Fee11', type=NUMBER)],
             [span('网元定制', row=2)],
             ['功能费', '套餐功能费', '-', put_input('Num12', type=NUMBER),
              '号', '6%',
              put_row([put_input('Discount12', type=NUMBER), put_text('%').style('margin-top: 5px')], size='80% 20%'),
              put_input('Fee12', type=NUMBER)]
             ]).style(
            'text-align:center;text-align-last:center;table-layout:fixed;word-wrap:break-word;width:1000px;margin:10px')

        # 更新表格
        while True:
            pin_wait_change('select1', 'Num1', 'Discount1', 'Fee1',
                            'select2', 'select3', 'Num2', 'Discount2', 'Fee2',
                            'select4', 'Num3', 'Discount3', 'Fee3',
                            'select5', 'Num5', 'Discount5',
                            'Num6', 'Discount6', 'Fee6',
                            'Num7', 'Discount7', 'Fee7',
                            'select11', 'Num11', 'Discount11', 'Fee11',
                            'Num12', 'Discount12', 'Fee12'
                            )

            # 查数据表，得到产品标准资费
            # 人联卡的产品标准资费
            with use_scope('scope2', clear=True):  # 删除scope1中的所有信息，重新建立表格
                # 人联卡相关信息计算
                a1 = Cal_RenLian(RenLian)
                # 数量
                num1 = a1[0]
                # 产品标准资费
                charge1 = a1[1]
                # 折扣率
                discount1 = a1[2]
                # 折后资费
                discountFee1 = a1[3]
                # 折后月资费小计
                discountMonthFee1 = a1[4]
                # 折后年资费合计
                discountYearFee1 = a1[5]
                # 一次性服务费/调试费
                oneTimeFee1 = a1[6]
                # 一次性服务费合计
                oneTimeFeeSum1 = a1[7]
                # 报价合计
                sum1 = a1[8]

                # 物联卡相关信息计算
                # 人联卡相关信息计算
                a2 = Cal_WuLian(WuLian_Month, WuLian_Year)
                # 数量
                num2 = a2[0]
                # 产品标准资费
                charge2 = a2[1]
                # 折扣率
                discount2 = a2[2]
                # 折后资费
                discountFee2 = a2[3]
                # 折后月资费小计
                discountMonthFee2 = a2[4]
                # 折后年资费合计
                discountYearFee2 = a2[5]
                # 一次性服务费/调试费
                oneTimeFee2 = a2[6]
                # 一次性服务费合计
                oneTimeFeeSum2 = a2[7]
                # 报价合计
                sum2 = a2[8]

                # 流量池相关信息计算
                a3 = Cal_LiuLiangChi(LiuLiangChi)
                # 数量
                num3 = a3[0]
                # 产品标准资费
                charge3 = a3[1]
                # 折扣率
                discount3 = a3[2]
                # 折后资费
                discountFee3 = a3[3]
                # 折后月资费小计
                discountMonthFee3 = a3[4]
                # 折后年资费合计
                discountYearFee3 = a3[5]
                # 一次性服务费/调试费
                oneTimeFee3 = a3[6]
                # 一次性服务费合计
                oneTimeFeeSum3 = a3[7]
                # 报价合计
                sum3 = a3[8]

                # 定制号卡相关信息计算
                a5 = Cal_HaoKa(HaoKa)
                # 数量
                num5 = a5[0]
                # 折扣率
                discount5 = a5[1]
                # 一次性服务费/调试费
                oneTimeFee5 = a5[2]
                # 一次性服务费合计
                oneTimeFeeSum5 = a5[3]
                # 报价合计
                sum5 = a5[4]

                # 定制DNN相关信息计算
                a6 = Cal_DNN(YeWuGeLi)
                # 数量
                num6 = a6[0]
                # 产品标准资费
                charge6 = a6[1]
                # 折扣率
                discount6 = a6[2]
                # 折后资费
                discountFee6 = a6[3]
                # 折后月资费小计
                discountMonthFee6 = a6[4]
                # 折后年资费合计
                discountYearFee6 = a6[5]
                # 一次性服务费/调试费
                oneTimeFee6 = a6[6]
                # 一次性服务费合计
                oneTimeFeeSum6 = a6[7]
                # 报价合计
                sum6 = a6[8]

                # 无线VPDN相关信息计算
                a7 = Cal_DNN(YeWuGeLi)
                # 数量
                num7 = a7[0]
                # 产品标准资费
                charge7 = a7[1]
                # 折扣率
                discount7 = a7[2]
                # 折后资费
                discountFee7 = a7[3]
                # 折后月资费小计
                discountMonthFee7 = a7[4]
                # 折后年资费合计
                discountYearFee7 = a7[5]
                # 一次性服务费/调试费
                oneTimeFee7 = a7[6]
                # 一次性服务费合计
                oneTimeFeeSum7 = a7[7]
                # 报价合计
                sum7 = a7[8]

                # 5G定制网IPRAN专线相关信息计算
                a11 = cal_IPRAN(IPRAN)
                # 数量
                num11 = a11[0]
                # 产品标准资费
                charge11 = a11[1]
                # 折扣率
                discount11 = a11[2]
                # 折后资费
                discountFee11 = a11[3]
                # 折后月资费小计
                discountMonthFee11 = a11[4]
                # 折后年资费合计
                discountYearFee11 = a11[5]
                # 一次性服务费/调试费
                oneTimeFee11 = a11[6]
                # 一次性服务费合计
                oneTimeFeeSum11 = a11[7]
                # 报价合计
                sum11 = a11[8]

                # 功能费相关
                a12 = cal_GonNeng(IPRAN)
                # 数量
                num12 = a12[0]
                # 产品标准资费
                charge12 = a12[1]
                # 折扣率
                discount12 = a12[2]
                # 折后资费
                discountFee12 = a12[3]
                # 折后月资费小计
                discountMonthFee12 = a12[4]
                # 折后年资费合计
                discountYearFee12 = a12[5]
                # 一次性服务费/调试费
                oneTimeFee12 = a12[6]
                # 一次性服务费合计
                oneTimeFeeSum12 = a12[7]
                # 报价合计
                sum12 = a12[8]

                # 计算总价
                YearSumList = [discountYearFee1, discountYearFee2, discountYearFee3, discountYearFee6, discountYearFee7,
                               discountYearFee11, discountYearFee12]
                OneTimeSumList = [oneTimeFeeSum1, oneTimeFeeSum2, oneTimeFeeSum3, oneTimeFeeSum5,
                                  oneTimeFeeSum6, oneTimeFeeSum7, oneTimeFeeSum11, oneTimeFeeSum12]
                TotalSumList = [sum1, sum2, sum3, sum5, sum6, sum7, sum11, sum12]
                YearSum = totalSum(YearSumList)
                OneTimeSum = totalSum(OneTimeSumList)
                TotalSum = totalSum(TotalSumList)

                put_text('5G VPDN资费报价').style('font-size:20px;font-weight:700;margin:15px')
                put_table(
                    [['', '名称', '类型', '规格', '数量', '单位', '税率', '产品标准资费', '折扣率', '折后资费',
                      '折后月资费小计', '折后年资费合计', '一次性服务费/调试费', '一次性成本合计', '报价合计'],
                     [span('定制流量', row=5)],
                     ['人联卡', '5G畅享套餐', put_text(" %s " % (pin.select1)), put_text(" %s " % (num1)), '张',
                      '6%', put_text(" %s " % (charge1)), put_text(" %s%s " % (discount1, "%")),
                      put_text(" %s " % (discountFee1)),
                      put_text(" %s " % (discountMonthFee1)), put_text(" %s " % (discountYearFee1)),
                      put_text(" %s " % (oneTimeFee1)), put_text(" %s " % (oneTimeFeeSum1)), sum1],
                     ['物联卡', put_text(" %s " % (pin.select2)), put_text(" %s " % (pin.select3)),
                      put_text(" %s " % (num2)), '张',
                      '6%', put_text(" %s " % (charge2)), put_text(" %s%s " % (discount2, "%")),
                      put_text(" %s " % (discountFee2)),
                      put_text(" %s " % (discountMonthFee2)), put_text(" %s " % (discountYearFee2)),
                      put_text(" %s " % (oneTimeFee2)), put_text(" %s " % (oneTimeFeeSum2)), sum2],
                     ['流量池', put_text(" %s " % (pin.select4)), '-',
                      put_text(" %s " % (num3)), 'GB',
                      '6%', put_text(" %s " % (charge3)), put_text(" %s%s " % (discount3, "%")),
                      put_text(" %s " % (discountFee3)),
                      put_text(" %s " % (discountMonthFee3)), put_text(" %s " % (discountYearFee3)),
                      put_text(" %s " % (oneTimeFee3)), put_text(" %s " % (oneTimeFeeSum3)), sum3],
                     ['定制号卡', put_text(" %s " % (pin.select5)), '-',
                      put_text(" %s " % (num5)), '张',
                      '一半6%,一半13%', '-', put_text(" %s%s " % (discount5, "%")),
                      '-', '-', '-',
                      put_text(" %s " % (oneTimeFee5)), put_text(" %s " % (oneTimeFeeSum5)), sum5],
                     [span('业务隔离', row=3)],
                     ['定制DNN', '专网加密隧道服务', '-',
                      put_text(" %s " % (num6)), '线',
                      '6%', put_text(" %s " % (charge6)), put_text(" %s%s " % (discount6, "%")),
                      put_text(" %s " % (discountFee6)),
                      put_text(" %s " % (discountMonthFee6)), put_text(" %s " % (discountYearFee6)),
                      put_text(" %s " % (oneTimeFee6)), put_text(" %s " % (oneTimeFeeSum6)), sum6],
                     ['无线 VPDN 群', '-', '-',
                      put_text(" %s " % (num7)), '线',
                      '6%', put_text(" %s " % (charge7)), put_text(" %s%s " % (discount7, "%")),
                      put_text(" %s " % (discountFee7)),
                      put_text(" %s " % (discountMonthFee7)), put_text(" %s " % (discountYearFee7)),
                      put_text(" %s " % (oneTimeFee7)), put_text(" %s " % (oneTimeFeeSum7)), sum7],
                     [span('固定入网专线', row=2)],
                     ['5G定制网IPRAN专线', put_text(" %s " % (pin.select11)), '-',
                      put_text(" %s " % (num11)), '线',
                      '6%', put_text(" %s " % (charge11)), put_text(" %s%s " % (discount11, "%")),
                      put_text(" %s " % (discountFee11)),
                      put_text(" %s " % (discountMonthFee11)), put_text(" %s " % (discountYearFee11)),
                      put_text(" %s " % (oneTimeFee11)), put_text(" %s " % (oneTimeFeeSum11)), sum11],
                     [span('网元定制', row=2)],
                     ['功能费', '套餐功能费', '-', put_text(" %s " % (num12)), '号',
                      '6%', put_text(" %s " % (charge12)), put_text(" %s%s " % (discount12, "%")),
                      put_text(" %s " % (discountFee12)),
                      put_text(" %s " % (discountMonthFee12)), put_text(" %s " % (discountYearFee12)),
                      put_text(" %s " % (oneTimeFee12)), put_text(" %s " % (oneTimeFeeSum12)), sum12],
                     ['项目总计', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', YearSum, '-', OneTimeSum, TotalSum]
                     ]).style(
                    'text-align:center;text-align-last:center;table-layout:fixed;word-wrap:break-word;width:1500px;margin:10px')

            # 点击下载表格
            with use_scope('scope3', clear=True):
                num = ['', num1, num2, num3, num5, num6, num7, num11, num12]
                charge = ['', charge1, charge2, charge3, '', charge6, charge7, charge11, charge12]
                discount = ['', discount1, discount2, discount3, discount5, discount6, discount7, discount11,
                            discount12]
                discountFee = ['', discountFee1, discountFee2, discountFee3, '', discountFee6,
                               discountFee7, discountFee11, discountFee12]
                discountMonthFee = ['', discountMonthFee1, discountMonthFee2, discountMonthFee3, '',
                                    discountMonthFee6, discountMonthFee7, discountMonthFee11, discountMonthFee12]
                discountYearFee = ['', discountYearFee1, discountYearFee2, discountYearFee3, '',
                                   discountYearFee6, discountYearFee7, discountYearFee11, discountYearFee12]
                oneTimeFee = ['', oneTimeFee1, oneTimeFee2, oneTimeFee3, oneTimeFee5, oneTimeFee6,
                              oneTimeFee7, oneTimeFee11, oneTimeFee12]
                oneTimeFeeSum = ['', oneTimeFeeSum1, oneTimeFeeSum2, oneTimeFeeSum3, oneTimeFeeSum5,
                                 oneTimeFeeSum6, oneTimeFeeSum7, oneTimeFeeSum11, oneTimeFeeSum12]
                sum = ['', sum1, sum2, sum3, sum5, sum6, sum7, sum11, sum12]
                put_button('报表下载',
                           onclick=lambda: VPDNFileDownload(num, charge, discount, discountFee, discountMonthFee,
                                                            discountYearFee, oneTimeFee, oneTimeFeeSum, sum,
                                                            YearSum, OneTimeSum, TotalSum)).style('margin:10px')


# 将更新后的信息填入表中，以供下载
def VPDNFileDownload(num, charge, discount, discountFee, discountMonthFee, discountYearFee, oneTimeFee,
                     oneTimeFeeSum, sum, YearSum, OneTimeSum, TotalSum):
    # 打开空白表格
    wb = openpyxl.load_workbook('./pywebio1/表格下载/5G_VPDN.xlsx')
    ws = wb.active
    # 人联卡相关
    ws.cell(row=6, column=5).value = "%s" % (pin.select1)
    ws.cell(row=6, column=6).value = "%s" % (num[1])
    ws.cell(row=6, column=9).value = "%s" % (charge[1])
    ws.cell(row=6, column=10).value = "%s%s" % (discount[1], "%")
    ws.cell(row=6, column=11).value = "%s" % (discountFee[1])
    ws.cell(row=6, column=12).value = "%s" % (discountMonthFee[1])
    ws.cell(row=6, column=13).value = "%s" % (discountYearFee[1])
    ws.cell(row=6, column=14).value = "%s" % (oneTimeFee[1])
    ws.cell(row=6, column=15).value = "%s" % (oneTimeFeeSum[1])
    ws.cell(row=6, column=16).value = "%s" % (sum[1])
    # 物联卡相关
    str = "%s" % (pin.select2)
    LeiXin = str[0:2]
    BeiZhu = str[3:5]
    ws.cell(row=7, column=3).value = "%s" % (LeiXin)
    ws.cell(row=7, column=4).value = "%s" % (BeiZhu)
    ws.cell(row=7, column=5).value = "%s" % (pin.select3)
    ws.cell(row=7, column=6).value = "%s" % (num[2])
    ws.cell(row=7, column=9).value = "%s" % (charge[2])
    ws.cell(row=7, column=10).value = "%s%s" % (discount[2], "%")
    ws.cell(row=7, column=11).value = "%s" % (discountFee[2])
    ws.cell(row=7, column=12).value = "%s" % (discountMonthFee[2])
    ws.cell(row=7, column=13).value = "%s" % (discountYearFee[2])
    ws.cell(row=7, column=14).value = "%s" % (oneTimeFee[2])
    ws.cell(row=7, column=15).value = "%s" % (oneTimeFeeSum[2])
    ws.cell(row=7, column=16).value = "%s" % (sum[2])
    # 流量池相关
    ws.cell(row=8, column=3).value = "%s" % (pin.select4)
    ws.cell(row=8, column=6).value = "%s" % (num[3])
    ws.cell(row=8, column=9).value = "%s" % (charge[3])
    ws.cell(row=8, column=10).value = "%s%s" % (discount[3], "%")
    ws.cell(row=8, column=11).value = "%s" % (discountFee[3])
    ws.cell(row=8, column=12).value = "%s" % (discountMonthFee[3])
    ws.cell(row=8, column=13).value = "%s" % (discountYearFee[3])
    ws.cell(row=8, column=14).value = "%s" % (oneTimeFee[3])
    ws.cell(row=8, column=15).value = "%s" % (oneTimeFeeSum[3])
    ws.cell(row=8, column=16).value = "%s" % (sum[3])
    # 定制号卡相关
    ws.cell(row=9, column=3).value = "%s" % (pin.select5)
    ws.cell(row=9, column=6).value = "%s" % (num[4])
    ws.cell(row=9, column=10).value = "%s%s" % (discount[4], "%")
    ws.cell(row=9, column=14).value = "%s" % (oneTimeFee[4])
    ws.cell(row=9, column=15).value = "%s" % (oneTimeFeeSum[4])
    ws.cell(row=9, column=16).value = "%s" % (sum[4])
    # 定制DNN相关
    ws.cell(row=11, column=6).value = "%s" % (num[5])
    ws.cell(row=11, column=9).value = "%s" % (charge[5])
    ws.cell(row=11, column=10).value = "%s%s" % (discount[5], "%")
    ws.cell(row=11, column=11).value = "%s" % (discountFee[5])
    ws.cell(row=11, column=12).value = "%s" % (discountMonthFee[5])
    ws.cell(row=11, column=13).value = "%s" % (discountYearFee[5])
    ws.cell(row=11, column=14).value = "%s" % (oneTimeFee[5])
    ws.cell(row=11, column=15).value = "%s" % (oneTimeFeeSum[5])
    ws.cell(row=11, column=16).value = "%s" % (sum[5])
    # 无线VPDN相关
    ws.cell(row=12, column=6).value = "%s" % (num[6])
    ws.cell(row=12, column=9).value = "%s" % (charge[6])
    ws.cell(row=12, column=10).value = "%s%s" % (discount[6], "%")
    ws.cell(row=12, column=11).value = "%s" % (discountFee[6])
    ws.cell(row=12, column=12).value = "%s" % (discountMonthFee[6])
    ws.cell(row=12, column=13).value = "%s" % (discountYearFee[6])
    ws.cell(row=12, column=14).value = "%s" % (oneTimeFee[6])
    ws.cell(row=12, column=15).value = "%s" % (oneTimeFeeSum[6])
    ws.cell(row=12, column=16).value = "%s" % (sum[6])
    # IPRAN专线
    ws.cell(row=14, column=3).value = "%s" % (pin.select11)
    ws.cell(row=14, column=6).value = "%s" % (num[7])
    ws.cell(row=14, column=9).value = "%s" % (charge[7])
    ws.cell(row=14, column=10).value = "%s%s" % (discount[7], "%")
    ws.cell(row=14, column=11).value = "%s" % (discountFee[7])
    ws.cell(row=14, column=12).value = "%s" % (discountMonthFee[7])
    ws.cell(row=14, column=13).value = "%s" % (discountYearFee[7])
    ws.cell(row=14, column=14).value = "%s" % (oneTimeFee[7])
    ws.cell(row=14, column=15).value = "%s" % (oneTimeFeeSum[7])
    ws.cell(row=14, column=16).value = "%s" % (sum[7])
    # 功能费
    ws.cell(row=16, column=3).value = "%s" % (pin.select11)
    ws.cell(row=16, column=6).value = "%s" % (num[8])
    ws.cell(row=16, column=9).value = "%s" % (charge[8])
    ws.cell(row=16, column=10).value = "%s%s" % (discount[8], "%")
    ws.cell(row=16, column=11).value = "%s" % (discountFee[8])
    ws.cell(row=16, column=12).value = "%s" % (discountMonthFee[8])
    ws.cell(row=16, column=13).value = "%s" % (discountYearFee[8])
    ws.cell(row=16, column=14).value = "%s" % (oneTimeFee[8])
    ws.cell(row=16, column=15).value = "%s" % (oneTimeFeeSum[8])
    ws.cell(row=16, column=16).value = "%s" % (sum[8])
    # 项目总和
    ws.cell(row=17, column=13).value = "%s" % (YearSum)
    ws.cell(row=17, column=15).value = "%s" % (OneTimeSum)
    ws.cell(row=17, column=16).value = "%s" % (TotalSum)

    update_FileName = './pywebio1/更新后的表/update-5G_VPDN.xlsx'
    wb.save(update_FileName)
    # 删除不需要的行，并补齐序号
    lims_file1 = lims_file(update_FileName)
    lims_file1.delete_space()
    lims_file1.serial()
    lims_file1.xlsxFormat()
    with use_scope('scope4', clear=True):
        content = open(update_FileName, 'rb').read()
        put_file('5G_VPDN.xlsx', content, '5G_VPDN.xlsx')

